import requests
from openpyxl import load_workbook

# Load Excel file
file_path = r"Week-3\\Day-2\\data_drive_api\\users (4).xlsx"
workbook = load_workbook(file_path)
workbook.active=workbook['product']
sheet = workbook.active
#sheet = workbook["product"]
#print(sheet)

url = "https://jsonplaceholder.typicode.com/users"

for row in sheet.iter_rows(min_row=2, values_only=True):
    name, version,price=row
    print("\nExecuting API for:", name)

    payload = {
        "name": name,
        "version": version,
        "price": price
    }

    # Send POST request
    response = requests.post(url, json=payload)

    data = response.json()

    # Print response
    print("Status Code:", response.status_code)
    print("Response:", data)

    # Validations
    assert response.status_code == 201
    assert data["name"] == name

    print("Validation Passed")
